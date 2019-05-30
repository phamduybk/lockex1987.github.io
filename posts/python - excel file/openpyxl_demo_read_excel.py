# -*- coding: utf-8 -*-
import openpyxl
from datetime import datetime


DATETIME_PATTERN = '%d/%m/%Y %H:%M:%S'

path = "sample.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active

# Dòng thứ nhất là tiêu đề
# Đọc bắt đầu dòng 2 đến hết
# Chú ý phải cộng 1 vào sheet.max_row vì range(n) sẽ chỉ đến n-1
for row in range(2, sheet.max_row + 1):
    employee_code = sheet['B' + str(row)].value
    full_name = sheet['C' + str(row)].value
    start_date = sheet['D' + str(row)].value
    salary = sheet['E' + str(row)].value

    # + full_name
    print(employee_code + ", " + ", " + datetime.strftime(start_date, DATETIME_PATTERN) + ", " + str(salary))
