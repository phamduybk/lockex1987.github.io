# -*- coding: utf-8 -*-
import datetime
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl import load_workbook


#wb = Workbook()
wb = load_workbook(filename = 'input.xlsx')
#wb = openpyxl.Workbook()

ws1 = wb.active
#ws1.title = "First sheet"
for row in range(10, 20):
	for col in range(3, 15):
		ws1.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(1, 40):
	ws3.append(range(10))

ws4 = wb.create_sheet(title="Types")
# set date using a Python datetime
ws4['A1'] = datetime.datetime(2010, 7, 21)
print(ws4['A1'].number_format)
#'yyyy-mm-dd h:mm:ss'

# You can enable type inference on a case-by-case basis
wb.guess_types = True
# set percentage using a string followed by the percent sign
ws4['B1'] = '3.14%'
wb.guess_types = False
print(ws4['B1'].value)
#0.031400000000000004
print(ws4['B1'].number_format)
#'0%'

# add a simple formula
ws5 = wb.create_sheet(title="Formula")
ws5["A1"] = "=SUM(1, 1)"

ws6 = wb.create_sheet(title="Image")
ws6['A1'] = 'You should see three logos below'
# create an image
# add to worksheet and anchor next to cells
img = Image('logo.png')
ws6.add_image(img, 'A1')

ws7 = wb.create_sheet(title="Hide columns")
ws7.column_dimensions.group('A', 'D', hidden=True)

wb.save('output.xlsx')
#wb.save(filename='output.xlsx')
